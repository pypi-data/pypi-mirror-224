import pickle
import sys

import pandas as pd
import openpyxl
import os
import datetime
import copy
import configparser
import ast

from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from vika import Vika


class WeekReportBuild:
    """
    自动化周报生成器。
    可读取指定数据源（基于本项目提供的模板），可来自维格表（在线）、Excel表格（本地）
    """

    # 维格表相关参数
    VIKA_TOKEN = str
    VIKA_DATA_SHEET = str

    # 排序及基本查询条件
    NEW_ORDER = None
    CONDITION = str

    # 数据源（excel/vika)
    data_source = "vika"
    local_data_file_name = str

    # 模板及缓存文件名
    template_file_name = str
    cache_file_name = str
    cache_file_path: str
    output_file_name = str
    path = str

    # 配置文件路径
    conf_path = "conf/"

    # 表头信息
    titledata = None

    # 配置解析器
    config = None

    # 初始化变量
    def __init__(self):
        self.template = None
        self.start_column = None
        self.start_row = None
        self.worksheet = None

        config_file_path = os.path.join(self.conf_path, 'config.ini')
        self.config = configparser.ConfigParser()
        self.config.read(config_file_path, 'utf-8')

        self.data_source = self.config.get("global_config","data_source")
        self.local_data_file_name = self.config.get("file_config","local_data_file_name")
        self.template_file_name = self.config.get("file_config", "template_file_name")
        self.cache_file_name = self.config.get("file_config", "cache_file_name")
        self.output_file_name = self.config.get("file_config", "output_file_name")
        self.path = self.config.get("file_config", "path")
        self.cache_file_path = os.path.join(self.path, self.cache_file_name)

        # 解析JSON字符串为Python对象
        json_data = self.config.get("prj_info", "titledata")
        self.titledata = ast.literal_eval(json_data)

        self.VIKA_TOKEN = self.config.get("viga_config", "VIKA_TOKEN")
        self.VIKA_DATA_SHEET = self.config.get("viga_config", "VIKA_DATA_SHEET")
        self.NEW_ORDER = self.config.get("viga_config", "NEW_ORDER").split(",")
        self.CONDITION = self.config.get("viga_config", "CONDITION")
        print(f"加载配置文件:[{config_file_path}]完成。")

    def clear_cache(self):
        """
        清除序列化缓存的数据。（通过网络请示获取的数据，会生成一个缓存文件，默认为：cache.pkl
        """
        if os.path.exists(self.cache_file_path):
            os.remove(self.cache_file_path)
            print(f"缓存文件[{self.cache_file_path}]删除成功！")

    def get_jobs_data(self):
        """
        加载数据（如有缓存会优先加载缓存，如希望获取最新数据，可手动调用：clear_cache()方法）
        """
        if os.path.exists(self.cache_file_path):
            return self.load_cache_data()
        elif self.data_source == "viga":
            return self.get_jobs_data_from_vika(self.CONDITION)
        elif self.data_source == "excel":
            return self.get_jobs_data_from_local(self.CONDITION)
        else:
            print(f"未定义的类型:{self.data_source}")
            print("遇到异常，结束程序。")
            sys.exit(1)  # 非零状态码，异常退出

    # 使用vika提供的API获取数据
    def get_jobs_data_from_vika(self, condition):

        print(f'开始从vika_api加载数据.')
        vika = Vika(self.VIKA_TOKEN)
        # 通过 datasheetId 来指定要从哪张维格表操作数据。
        datasheet = vika.datasheet(self.VIKA_DATA_SHEET, field_key="name")

        # 返回满足条件的集合
        records = datasheet.records.filter(所属项目=condition, 周报="是")
        print(f'从vika_api成功加载数据:{len(records)}条.')
        fields_data = []
        for record in records:
            fields_data.append(record.json())

        # 将数据转换为pandas dataframe
        df = pd.DataFrame(fields_data)

        # 调整列顺序
        df = df.reindex(columns=self.NEW_ORDER)

        # 调整序号从1开始
        df['序号'] = df.index + 1

        # 缓存数据如果不存在，自动缓存数据
        if not os.path.exists(self.cache_file_path):
            self.cache_data(df)
        return df

    # 初始加载模板
    def load_template(self):

        # 打开Excel模板文件
        input_file_path = os.path.join(self.path, self.template_file_name)
        self.template = openpyxl.load_workbook(input_file_path)

        # 选择需要填写的工作表
        self.worksheet = self.template['周报主体']

        # 生成本周开始结束时间
        today = datetime.date.today()
        start_of_week = today - datetime.timedelta(days=today.weekday() + 1)
        end_of_week = start_of_week + datetime.timedelta(days=5)

        # 遍历单元格，替换占位符
        for row in self.worksheet:
            for cell in row:
                if cell.value and isinstance(cell.value, str) and '{' in cell.value and '}' in cell.value:

                    # 自动填写周报填写日期
                    if '{date}' in str(cell.value):
                        # 获取当前日期并格式化为"yyyy-MM-DD"的格式
                        now = datetime.date.today()
                        today = now.strftime('%Y-%m-%d')
                        # 替换单元格内容为当前日期
                        new_value = str(cell.value).replace('{date}', today)
                        cell.value = new_value

                    # 自动填写周报开始时间
                    elif '{start_of_week}' in str(cell.value):
                        new_value = str(cell.value).replace('{start_of_week}', str(start_of_week))
                        cell.value = new_value

                    # 自动填写周报结束时间
                    elif '{end_of_week}' in str(cell.value):
                        new_value = str(cell.value).replace('{end_of_week}', str(end_of_week))
                        cell.value = new_value

                    # 遍历数据字典，将点位符替换为对应的值
                    else:
                        for key, value in self.titledata.items():
                            if '{' + key + '}' in cell.value:
                                cell.value = cell.value.replace('{' + key + '}', value)

        # 找到工作任务表回填单元格位置
        for row in self.worksheet:
            for cell in row:
                if '{week_jobs}' in str(cell.value):
                    self.start_row = cell.row
                    self.start_column = cell.column
                    cell.value = ""
                    break
        print(f"加载模板:{self.template_file_name}完成！")

    # 回填具体工作事务信息
    def fill_work_data(self, df):

        # 回填第一列标题
        cell = self.worksheet.cell(row=self.start_row - 1, column=self.start_column)
        cell.value = "序号"
        col_names = df.columns
        bold_font = Font(bold=True)
        columns_count = len(col_names)
        print(f"累计数据列:{columns_count}")
        # 回填标题列，并设置粗体
        for n in range(columns_count):
            cell = self.worksheet.cell(row=self.start_row - 1, column=self.start_column + n)
            cell.value = col_names[n]
            cell.font = bold_font
        # 遍历DataFrame中的每一行，转换为行数据，并逐行追加到Excel表格中
        row_count = len(df)
        print(f"累计数据行：{row_count}")
        # 保留指定单元格样式
        source_cell = self.worksheet.cell(row=self.start_row, column=self.start_column)
        source_font = copy.copy(source_cell.font)
        source_fill = copy.copy(source_cell.fill)
        # 根据数据条数，插入若干空行
        self.worksheet.insert_rows(self.start_row, amount=row_count)
        # 逐行逐列回填数据
        # 表格边框
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))
        for i, row in df.iterrows():
            row_data = [row[col] for col in df.columns]
            # self.worksheet.row_dimensions[cell.row].auto_size = True  # 自动行高
            for j, cell_value in enumerate(row_data):
                cell = self.worksheet.cell(row=self.start_row + i, column=self.start_column + j)
                cell.value = cell_value
                cell.font = source_font # 字体
                cell.border = border # 边框
                cell.fill = source_fill # 单元格背景色
                cell.alignment = Alignment(wrap_text=True)# 自动换行

                # for merged_range in self.worksheet.merged_cells.ranges:
                #     if cell.coordinate in merged_range:
                #         self.worksheet.unmerge_cells(str(merged_range[0]))

    # 保存填写后的Excel文件
    def save_file(self):
        output_file_path = os.path.join(self.path, self.output_file_name)
        self.template.save(output_file_path)
        print(f'文件已输出至：{output_file_path}')

    # 缓存数据
    def cache_data(self, data):
        with open(self.cache_file_path, 'wb') as f:
            pickle.dump(data, f)
        print(f"数据缓存完成,文件：{self.cache_file_path}")

    # 加载缓存数据
    def load_cache_data(self):
        with open(self.cache_file_path, 'rb') as f:
            cached_data = pickle.load(f)
        print(f"从缓存[{self.cache_file_path}]加载数据：{len(cached_data)}")
        return cached_data

    # 从本地读取数据
    def get_jobs_data_from_local(self, condition):
        local_data_file_path = os.path.join(self.path, self.local_data_file_name)
        print(f"加载本地数据源：{local_data_file_path}")
        df_local = pd.read_excel(local_data_file_path)
        # 根据项目筛选数据
        df_local = df_local[df_local['所属项目'] == condition]
        # 调整列顺序
        df_local = df_local.reindex(columns=self.NEW_ORDER)
        # 调整序号从1开始
        df_local['序号'] = df_local.index + 1
        print('读取记录总数：', len(df_local))
        return df_local

if __name__ == '__main__':
    tool = WeekReportBuild()
    tool.load_template()
    tool.clear_cache()
    df = tool.get_jobs_data()
    tool.fill_work_data(df)
    tool.save_file()
