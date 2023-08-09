# -*- encoding: utf-8 -*-
"""
@File    :   write.py
@Time    :   2023-02-17 23:02
@Author  :   坐公交也用券
@Version :   1.0
@Contact :   faith01238@hotmail.com
@Homepage : https://liumou.site
@Desc    :   使用Openpyxl模块实现的写入功能
"""
from loguru import logger
from openpyxl import Workbook, load_workbook

from .exists import _exists


class Write:
	def __init__(self, filename):
		"""
		写入表格数据
		:param filename: 需要写入的表格文件
		"""
		self.logger = logger
		self.filename = filename
		if _exists(filename):
			self.wb = Workbook()
		else:
			self.wb = load_workbook(filename=self.filename)
		self.ws = self.wb.active
		self.SetSerialNumber = False  # 是否自动添加序号(默认False)
		self.SetAdd = True
		self.SetHead = True
		self.SetLine = None
		self.SetSpeedSum = 100
		self._WrData = []  # 设置需要写入的数据列表
		self.Err = None

	def create_sheet(self, name, index=None):
		"""
		创建sheet
		:param name: 新sheet的名称
		:param index: 设置sheet排序位置(0是排第一)
		:return: 创建结果
		"""
		if index is None:
			self.wb.create_sheet(title=name)
		else:
			self.wb.create_sheet(title=name, index=index)
		self.wb.save(self.filename)

	def UpdateLine(self, row, valueList):
		"""
		更新某行数据
		:param row: 需要更新的行
		:param valueList: 行数据列表
		:return:
		"""
		try:
			col = 0
			for i in valueList:
				self.ws.cell(row=row, column=col, value=i)
				col += 1
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def AddList(self, lists):
		"""
		通过列表方式写入数据,一次性最多写入104万行数据
		:param lists: 写入数据列表,例如: [["张三", "男", "33"], ["李四", "男", "32"]]
		:return:
		"""
		try:
			for i in lists:
				self.ws.append(i)
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def AddLine(self, data):
		"""
		追加写入一行数据
		:param data: 数据,以列表形式 ["张三", "男", "33"]
		或者字典模式1: {"A": "刘某", "B": "男", "C": "22"}
		字典模式2: {1: 1, 2: 2, 3: 3}
		:return:
		"""
		try:
			self.ws.append(data)
			self.wb.save(self.filename)
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def AddCol(self, col, data):
		"""
		写入一列数据
		:param col: 列数
		:param data: 数据
		:return:
		"""
		try:
			self.ws.insert_cols(idx=col, amount=1)
			# self.ws.
			return True
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
			return False

	def DeleteLine(self, index, row=1):
		"""
		删除行数据
		:param index: 需要删除的起始行
		:param row: 需要删除的行数总数,默认删除1行(也就是起始行)
		:return: 删除结果
		"""
		self.Err = None
		try:
			self.ws.delete_rows(idx=index, amount=row)
			self.wb.save(self.filename)
		except Exception as err:
			self.Err = err
			self.logger.error(self.Err)
		return self

	def Set(self, add=True, head=True, line=None, speed_sum=100, serial_number=False):
		"""
		设置写入参数
		:param add: 是否使用追加模式(默认: True)
		:param head: 是否保留表头标题(默认: True)
		:param line: 是否自定义写入的行，如果需要自定义，请传所在行的整数
		:param speed_sum: 写入多少行数据进行一次进度显示
		:param serial_number: 是否自动添加序号(默认False)
		:return:
		"""
		self.SetSerialNumber = serial_number
		self.SetAdd = add
		self.SetHead = head
		self.SetLine = line
		if not self.SetAdd:
			# 如果使用覆盖模式,则设置初始行: 0
			self.SetLine = 0
		self.SetSpeedSum = speed_sum
