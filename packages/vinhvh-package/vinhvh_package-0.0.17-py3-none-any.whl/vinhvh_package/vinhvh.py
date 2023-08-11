
import PySimpleGUI as sg
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import time
import os
from pathlib import Path
import logging
from pykml import parser
import sys
from geopy.distance import geodesic
import simplekml


def select_file(file: list = None, value: list = None, option: dict = None, checkbox: list = None, mode: str = 'single', title: str = 'File Selector'):
    """
    This function create a interface for user with different option input

    Parameters
    ----------
    file: (list, optional)
        File to browse
    value: (list, optional)
        Value for input
    option: (dict, optional)
        Combobox for select
    checkbox: (list, optional)
        Option for select
    tilte: (str, optional)
        Name of window
    mode: (str, optional) 'single' | 'multi'

    Return
    ----------
    List of select value

    """
    # Set the PySimpleGUI theme
    sg.theme('Dark Blue 3')

    # Create the layout for the file selector window
    layout = [[sg.Text('Select values ')]]
    file_paths = []
    len_path = 0
    if file is not None:
        if mode == 'single':
            for name in file:
                layout.append([sg.Text(name + ": ", size=(15, 1),),
                               sg.InputText(), sg.FileBrowse()])
        elif mode == 'multi':
            for name in file:
                layout.append([sg.Text(name + ": ", size=(15, 1),),
                               sg.InputText(key=f'{name}'), sg.FilesBrowse()])

    if value is not None:
        for name in value:
            layout.append([sg.Text(name + ": ", size=(15, 1),),
                           sg.InputText()])

    if option is not None:
        for name, value in option.items():
            layout.append([sg.Text(name + ': ', size=(15, 1)),
                          sg.Combo(value, default_value=value[0], size=(43, 1))],)

    if checkbox is not None:
        for name in checkbox:
            layout.append([sg.Checkbox(name + ':', default=False)])

    # Add the Submit and Cancel buttons to the layout
    layout.append([sg.Submit(), sg.Cancel()])

    # Create the file selector window
    window = sg.Window(title, layout)

    # Wait for the user to interact with the window
    event, values = window.read()

    # Close the window
    window.close()

    # Process the user's selection
    if event == 'Submit':

        if file is not None:
            # If in single mode, get the file paths from the input fields
            if mode == 'single':
                path = [values[i] for i in range(len(file))]
                len_path += len(file)

            # If in multi mode, get the file paths as a list of lists from the input fields
            elif mode == 'multi':
                path = [values[f'{name}'].split(';') for name in file]
            file_paths += path

        if value is not None:
            path = [values[i] for i in range(len_path, len_path + len(value))]
            file_paths += path
            len_path += len(value)

        if option is not None:
            path = [values[i] for i in range(len_path, len_path + len(option))]
            file_paths += path
            len_path += len(option)

        if checkbox is not None:
            path = [values[i]
                    for i in range(len_path, len_path + len(checkbox))]
            file_paths += path
            len_path += len(checkbox)

            # Return the list of file paths
        return file_paths

    # If the user clicked the Cancel button, print a message and exit the program
    elif event == 'Cancel':
        print('File selection cancelled.')
        sys.exit()
    # If the user closed the window, print a message and exit the program
    elif event == sg.WIN_CLOSED:
        print('File selection closed.')
        sys.exit()
    # If the user didn't select any files, return None
    else:
        return None


def read_file(filepath: str, sheet_name: str = None, skip_row: int = 0, sep: str = ','):
    """
    This function read form filepath and return to dataframe

    Parameters
    ----------
    file: (str)
        File to browse
    sheet_name: (str, optional)
        Name of sheet to read
    skip_row: (st, optional)
        Number of first row to skip
    sep: (str, optional)
        sep of csv file
    Return
    ----------
    Dataframe
    """
    # Determine file type based on extension
    if filepath.lower().endswith(('.txt', '.csv', '.xlsx', '.xls', '.log', '.kml')):
        if filepath.lower().endswith('.txt'):
            df = pd.read_csv(filepath, sep='	',
                             skiprows=skip_row, low_memory=False)
        elif filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath, sep=sep, skiprows=skip_row,
                             low_memory=False)
        elif filepath.lower().endswith('.xlsx'):
            if sheet_name is None:
                df = pd.read_excel(
                    filepath, skiprows=skip_row, engine='openpyxl')
            else:
                df = pd.read_excel(
                    filepath, sheet_name=sheet_name, skiprows=skip_row, engine='openpyxl')
        elif filepath.lower().endswith('.xls'):
            if sheet_name is None:
                df = pd.read_excel(filepath, skiprows=skip_row, engine='xlrd')
            else:
                df = pd.read_excel(
                    filepath, sheet_name=sheet_name, skiprows=skip_row, engine='xlrd')
        elif filepath.lower().endswith('.log'):
            with open(filepath, 'r') as f:
                lines = f.readlines()
            df = pd.DataFrame(lines, columns=['log'])
        elif filepath.lower().endswith('.kml'):
            # Parse the KML file using the pykml parser
            with open(filepath) as f:
                root = parser.parse(f).getroot()

            # Access the elements in the KML file
            document = root.Document
            data = []
            for placemark in document.Placemark:
                name = placemark.name.text
                latitude = placemark.Point.coordinates.text.split(',')[1]
                longitude = placemark.Point.coordinates.text.split(',')[0]
                data.append({'Name': name, 'Lat': latitude,
                            'Long': longitude})

            # Convert data to DataFrame
            df = pd.DataFrame(data, columns=['Name', 'Lat', 'Long'])

            # Convert Lat Long to float
            df['Lat'] = df['Lat'].astype(float)
            df['Long'] = df['Long'].astype(float)

        return df
    else:
        raise ValueError("Unsupported file format")


def save_file(df_result: pd.DataFrame, file_name: str = 'Program running result', folder: str = "Result", type: str = 'excel'):
    """
    This function saves a DataFrame to a file.

    Parameters
    ----------
    df_result : pandas.DataFrame
        The DataFrame to save.
    file_name : str, optional
        The name of the file to save.
    folder : str, optional
        The name of the folder to save the file in.
    file_type : str, optional
        The type of file to save the DataFrame as. Can be one of 'excel', 'txt', 'csv', 'kml'.


    Returns
    -------
    str
        The full path of the saved file.
    """

    # Generate a timestamp for the current date and time
    date_time = f"{datetime.now():%d%m%y_%H%M%S}"

    # Create the result directory if it doesn't exist
    os.makedirs(folder, exist_ok=True)

    # Change the current working directory to the result directory
    os.chdir(folder)

    if type == 'excel':
        file_name = f'{file_name} {date_time}.xlsx'
        df_result.to_excel(file_name, index=False)
    elif type == 'txt':
        file_name = f'{file_name} {date_time}.txt'
        df_result.to_csv(file_name, sep='\t', index=False)
    elif type == 'csv':
        file_name = f'{file_name} {date_time}.csv'
        df_result.to_csv(file_name, index=False)
    elif type == 'kml':
        df_result.save(file_name)
    else:
        file_path = None
        print('Does not support file type for saving!')

    # Get the full file path of the saved file
    file_path = os.path.abspath(file_name)

    # Change the current working directory back to the main source directory
    os.chdir(os.path.dirname(os.getcwd()))

    print('File saved successfully!')

    # Return the full file path of the saved file
    return file_path


def format_excel(file_name: str, column_width: dict = None):
    """
    This function format excel file to look more beautiful

    Parameters
    ----------
    file_name: (str)
        Name of excel file need to format
    comlumn_width: (dict, optional)
        width of column want to change

    return
    ----------
    None
    """

    # Load the Excel spreadsheet
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Set the font and alignment styles for all cells
    font = Font(name='Calibri', size=11)
    alignment = Alignment(horizontal='center',
                          vertical='center')

    # Add borders to all cells
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    # Set the background color and font and alignment styles for the first row
    fill = PatternFill(start_color='FFC000',
                       end_color='FFC000', fill_type='solid')
    font_first_row = Font(name='Calibri', size=11, bold=True)
    alignment_first_row = Alignment(
        horizontal='center', vertical='center', wrap_text=True)

    # Set all columns to the same width as the original file
    for i, column in enumerate(ws.columns):
        ws.column_dimensions[column[0].column_letter].width = 15
    if column_width is not None:
        for column, width in column_width.items():
            ws.column_dimensions[column].width = int(width)

    # Format all cells in the worksheet
    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                # Format the first row cells
                cell.font = font_first_row
                cell.alignment = alignment_first_row
                cell.border = border
                cell.fill = fill
            else:
                # Format all other cells
                cell.font = font
                cell.alignment = alignment
                cell.border = border

    # Save the modified Excel spreadsheet
    wb.save(file_name)
    print('Finish format excel file!')
    return file_name


def check_license(start_date_str: str, duration_days: int = 90):
    """
    This function create license by day

    Parameters
    ----------
    start_date_str: (str)
        start day: 'YYYY-MM-DD'
    duration_days: (int)
        Day of license
    return
    ----------
    Boolen
    """
    # Convert the start date string to a datetime object
    start_date = datetime.fromisoformat(start_date_str)

    # Calculate the end date based on the start date and duration in days
    end_date = start_date + timedelta(days=duration_days)

    # Get the current date and time
    now_date = datetime.now()

    # Calculate the time difference between the end date and the current date in seconds
    time_compare = end_date.timestamp() - now_date.timestamp()

    # Calculate the number of whole days remaining in the license period
    days_remaining = int(time_compare / 86400)

    # Determine the status of the license based on the number of days remaining
    if days_remaining < 0:
        print('Your program license has been expired, please contact to provider for more information!')
        return False
        time.sleep(10)
    elif days_remaining == 0 or days_remaining == 1:
        print(f'License: {days_remaining} day')
        return True
    elif days_remaining <= 30 and days_remaining > 1:
        print('Lic = True')
        print(f'License: {days_remaining} days')
        return True
    else:
        print('Lic = True')
        return True


def popup_finish(program_running_time, result_path: str):
    """
    This function popup finish message

    Parameters
    ----------
    program_running_time: 
        running time duration
    result_path: (str)
        DPath to result file
    return
    ----------
    None
    """
    # Create the message to display in the popup window
    message = '        Finish!' + '            ' + '\n' +\
        '        Program running time: ' + program_running_time + '            ' + '\n' +\
        '        Result: ' + result_path + '\n' +\
        '        Copyright: VinhVH      '

    # Display the popup window with the message and a title
    sg.popup_auto_close(
        message, title='Program running result', auto_close_duration=20)


def popup_error():
    """
    This function popup error message and save error logfile in Logfile folder

    Parameters
    ----------
    return
    ----------
    None
    """
    # Create the result directory if it doesn't exist
    Path("Logfile").mkdir(exist_ok=True)
    # Change the current working directory to the Logfile directory
    os.chdir("Logfile")
    date_time = datetime.now().strftime("%d%m%y_%H%M%S")
    logfile = f'Error logfile {date_time}.log'

    # Configure the logging module
    logging.basicConfig(filename=logfile, level=logging.ERROR)

    try:
        # Your code that may raise an exception
        raise Exception("Something went wrong")
    except Exception as e:
        # Log the exception to the file
        logging.exception("An exception was thrown: %s", str(e))
        # Print the exception to the terminal
        print("An exception was thrown: ", str(e))
        # Print the exception traceback to the terminal
        exc_type, exc_value, exc_traceback = sys.exc_info()
        traceback_details = {
            'filename': exc_traceback.tb_frame.f_code.co_filename,
            'lineno': exc_traceback.tb_lineno,
            'name': exc_traceback.tb_frame.f_code.co_name,
            'type': exc_type.__name__,
            'message': str(exc_value)
        }
        for key, value in traceback_details.items():
            print(f"{key}: {value}")

    message = f'        Error \n        Bye bye!'
    sg.popup_auto_close(message, title='Notice',
                        auto_close_duration=20, background_color='red')
    time.sleep(10)


def start():
    st_time = time.time()
    start_time = datetime.now()
    print('Program is running...')
    print('Start running time:', start_time.strftime("%d/%m/%Y, %H:%M:%S"))
    return st_time


def end(st_time, result_path: str):
    end_time = time.time()
    finish_time = datetime.now()
    print('End running time:', finish_time.strftime("%d/%m/%Y, %H:%M:%S"))
    running_time = end_time - st_time
    program_running_time = f"{int(running_time//60)}'" + \
        f'{int(running_time%60)}s'
    print('Program running time:', program_running_time)
    print('Mission completed!')
    message = f'        Finish!\n        Program running time: {program_running_time}\n        Result path: {result_path}\n        Copyright: VinhVH'
    sg.popup_auto_close(
        message, title='Running result', auto_close_duration=20)


def the_most_common_value(df: pd.DataFrame):
    """
    This function create the most common value by parameter

    Parameters
    ----------
    df: (Dataframe)

    return
    ----------
    Dict of common value
    """
    common_values = {}
    for col in df.columns:
        mode = df[col].mode()
        if not mode.empty:
            common_values[col] = mode.iloc[0]
        else:
            common_values[col] = 'Empty'
    return common_values


def type_site(x: str):
    """
    This function create the most common value by parameter

    Parameters
    ----------
    x: (str)
    str to check

    return
    ----------
    Dict of common value
    """

    site_types = {
        'IBC': ['7', '8', '9', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'I4BA', 'I4BB', 'I4BC', 'I4BD', 'I4BE', 'I4BF', 'I4BJ', 'I4BK', 'I4CA', 'I4CB', 'I4CC', 'I4CD', 'I4CE', 'I4CF', 'I4CJ', 'I4CK'],
        'Macro': ['COG', 'COH', 'COI', '1', '2', '3', '4', '5', '6', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'M4BA', 'M4BB', 'M4BC', 'M4BD', 'M4BE', 'M4BF', 'M4BG', 'M4BH', 'M4BI', 'M4CA', 'M4CB', 'M4CC', 'M4CD', 'M4CE', 'U4BA', 'U4BB', 'U4BC', 'U4CA', 'U4CB', 'U4CC', 'A_LTE', 'B_LTE', 'C_LTE', 'D_LTE', 'E_LTE', 'F_LTE', '_LTE', 'M4BF', 'M4CF', 'M4RA', 'M4RB', 'M4RC', 'U4BA', 'U4BB', 'U4BC'],
        'CRAN': ['J4BA', 'J4BB', 'J4BC', 'J4CA', 'J4CB', 'J4CC', 'C4BA', 'C4BB', 'C4BC', 'C4CA', 'C4CB', 'C4CC'],
        'Smallcell': ['S', 'T'],
        'S4': ['W', 'V', 'X', 'U', 'Z'],
        'IoT': ['_IoT', '_IOT']
    }
    for site_type, codes in site_types.items():
        if x in codes:
            return site_type
    return 'Other'


def dvt(x: str):
    """
    This function create the most common value by parameter

    Parameters
    ----------
    x: (str)
    str to check

    return
    ----------
    Name of DVT
    """

    site_types = {
        'DVTD': ['H01', 'H02', 'H03', 'H04', 'H05', 'H07', 'H09', 'H10', 'H11', 'HBT', 'HCG', 'HNB', 'HPN', 'HTD'],
        'DVTT': ['H06', 'H08', 'H12', 'HBC', 'HBI', 'HCC', 'HGV', 'HHM', 'HTB', 'HTP']
    }
    for site_type, codes in site_types.items():
        if x in codes:
            return site_type
    return 'Other'


def date_time():
    return f"{datetime.now():%d%m%y_%H%M%S}"


def calculate_distance(lat1, lon1, lat2, lon2):
    # Define the coordinates of the two points as tuples
    point1 = (lat1, lon1)
    point2 = (lat2, lon2)

    # Calculate the distance between the points using the geodesic function
    distance = geodesic(point1, point2).km

    return distance


def make_kml_file(df_data: pd.DataFrame, date_time: str = date_time()):

    # Create site_dict
    site_dict = df_data.set_index(
        'Site')[['Lat', 'Long', 'Area']].to_dict(orient='index')

    # Create a KML object
    kml = simplekml.Kml()

    # Define colors for the different classes
    cran_color = 'ff00ff00'  # green
    main_color = 'ff0000ff'  # red
    line_color = 'ff00ffff'  # yellow

    # Define icons for the different classes
    cran_icon = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'
    main_icon = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'

    # Loop through the rows of the dataframe and add a point for each site
    for index, row in df_data.iterrows():
        # Get the coordinates of the site
        lon, lat = row['Long'], row['Lat']

        # Define the color and size of the point based on the class
        if row['Class'] == 'Main':
            color = main_color
            icon = main_icon
            size = 1.7
        else:
            color = cran_color
            icon = cran_icon
            size = 1.2

        # Create a point object for the site
        pnt = kml.newpoint()
        pnt.name = f"{row['Site']}"
        pnt.description = f"Class: {row['Class']}\nGroup code: {row['Group code']}"
        pnt.coords = [(lon, lat)]
        pnt.style.iconstyle.icon.href = icon
        pnt.style.iconstyle.color = color
        pnt.style.iconstyle.scale = size

        # If the site is not the main site, draw a line to the main site
        main_site = row['Main site']
        main_lat = site_dict[main_site]['Lat']
        main_lon = site_dict[main_site]['Long']

        if row['Class'] != 'Main':
            line = kml.newlinestring()
            line.coords = [(lon, lat), (main_lon, main_lat)]
            line.style.linestyle.color = line_color
            line.style.linestyle.width = 1.5
    # Luu ket qua ra file excel trong thu muc result
    result_dir = Path('Result')
    result_dir.mkdir(exist_ok=True)
    os.chdir(result_dir)

    # Save the KML file
    file_name = f"CRAN checking result {date_time}.kml"
    kml.save(file_name)

    print('Finish create kml file!')
    # Change the current working directory back to the main source directory
    os.chdir(os.path.dirname(os.getcwd()))
    return file_name
