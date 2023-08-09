"""
@author: lijc210@163.com
@file: 2.py
@time: 2019/11/27
@desc: 功能描述。
"""
import json
from collections import defaultdict

from openpyxl import load_workbook

d = {}
# 打开一个workbook
wb = load_workbook(filename="枚举数据.xlsx", data_only=True)

# 获取所有表格(worksheet)的名字
sheets = wb.sheetnames

for sheet in sheets:
    print(sheet)
    if sheet == "施工进度-投诉":
        continue
    # 获取特定的worksheet
    ws = wb[sheet]

    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns

    tmp_dict = defaultdict(list)

    row1 = []
    # 迭代所有的行
    for i, row in enumerate(rows, 1):
        if i == 1:
            row1 = [col.value for col in row]
        else:
            line = [col.value for col in row]
            if set(line) == {None}:
                continue
            adict = dict(zip(row1, line, strict=True))

            shop_name = adict.get("店铺名称")
            shop_id = adict.get("店铺id")
            if not shop_id:
                shop_id = adict.get("商家id")
                if not shop_id:
                    shop_id = adict.get("zx_shop_id")
                    if not shop_id:
                        raise ValueError("aaaaaaaaaaaaaaa")
            if not shop_name:
                shop_name = adict.get("商家名称")
                if not shop_name:
                    raise ValueError("bbbbbbbbbbbbb")
            tmp_dict[shop_name + "-" + str(shop_id)].append(adict)
    d[sheet] = tmp_dict

print(d["锁单接单结构-订单区域分布"]["天怡美装饰-215227980"])

with open("data.json", "w", encoding="utf-8") as f:
    f.write(json.dumps(d, ensure_ascii=False))
