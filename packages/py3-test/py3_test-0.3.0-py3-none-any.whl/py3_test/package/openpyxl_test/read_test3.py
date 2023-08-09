"""
@author: lijc210@163.com
@file: 2.py
@time: 2019/11/27
@desc: 功能描述。
"""
import json

from openpyxl import load_workbook

l1 = []
# 打开一个workbook
wb = load_workbook(filename="汇总数据.xlsx", data_only=True)

# 获取所有表格(worksheet)的名字
sheets = wb.sheetnames

for sheet in sheets:
    if sheet == "汇总":
        # 获取特定的worksheet
        ws = wb[sheet]

        # 获取表格所有行和列，两者都是可迭代的
        rows = ws.rows
        columns = ws.columns

        row1 = []
        # 迭代所有的行
        for i, row in enumerate(rows, 1):
            if i == 1:
                row1 = [col.value for col in row]
            else:
                line = [col.value for col in row]
                if set(line) == {None}:
                    continue
                adict = dict(zip(row1, line, strict=True))
                l1.append(adict)

with open("huizong.json", "w", encoding="utf-8") as f:
    f.write(json.dumps(l1, ensure_ascii=False))
